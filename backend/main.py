"""
NYC UAP / 485-x development strategy backend.
Uses Pinecone RAG + GPT multi-agent orchestration to evaluate zoning, site context, and developer-oriented scenarios.
"""

import os
import io
import json
import logging
from contextlib import asynccontextmanager

from dotenv import load_dotenv
from fastapi import FastAPI, UploadFile, File, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openai import OpenAI
from pinecone import Pinecone, ServerlessSpec
from pydantic import BaseModel

from engine.engine import context_engine
from engine.helpers import helper_sanitize_input, helper_moderate_content, get_embedding, get_embeddings_batch
from property_models import (
    BlockLotsResponse,
    PropertyContext,
    PropertyContextRequest,
    PropertySearchResponse,
    ValidatedLotInfo,
)
from property_service import property_service
from property_store import delete_property_context, fetch_property_context, upsert_property_context
from underwriting_calculator import (
    build_underwriting_calculation_context,
    calculate_underwriting_formula_values,
    enable_workbook_recalculation,
)
from underwriting_template import build_underwriting_cell_payload

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "..", ".env"))

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# ── Configuration ───────────────────────────────────────────────────────

GENERATION_MODEL = os.getenv("GENERATION_MODEL", "gpt-5.4")
EMBEDDING_MODEL = os.getenv("EMBEDDING_MODEL", "text-embedding-3-large")
PINECONE_INDEX = os.getenv("PINECONE_INDEX", "und1")
NAMESPACE_CONTEXT = os.getenv("NAMESPACE_CONTEXT", "ContextLibrary")
NAMESPACE_KNOWLEDGE = os.getenv("NAMESPACE_KNOWLEDGE", "KnowledgeStore")
NAMESPACE_PROPERTY = os.getenv("NAMESPACE_PROPERTY", "PropertyContextStore")
DEFAULT_CORS_ALLOW_ORIGINS = ("http://localhost:3000", "http://localhost:3001", "http://127.0.0.1:35095", "http://localhost:35095")


def _get_csv_env(name: str, default: tuple[str, ...]) -> list[str]:
    raw_value = os.getenv(name, "")
    if not raw_value.strip():
        return list(default)
    values = [value.strip() for value in raw_value.split(",") if value.strip()]
    return values or list(default)


import re as _re

_DOMAIN_KEYWORDS = _re.compile(
    r"uap|485[\-\s]?x|zoning|far\b|floor area ratio|underwriting|rent roll|"
    r"t[\-\s]?12|noi\b|cap rate|debt service|dscr|affordable housing|"
    r"ami\b|hpd|hdc|nyc housing|tax abatement|tax exemption|"
    r"operating (expenses?|statement)|pro ?forma|rent stabiliz|"
    r"offering memorandum|appraisal|bbl\b|borough|lot area|building area|"
    r"development (strategy|scenario|site)|buildable|residential far|"
    r"commercial far|community facility",
    _re.IGNORECASE,
)

_SITE_KEYWORDS = _re.compile(
    r"\b(this|the|our|my)\s+(site|property|building|parcel|lot|project)\b",
    _re.IGNORECASE,
)


def _is_domain_query(query: str, context) -> bool:
    """Return True if *query* is within the NYC UAP / 485-x domain."""
    if _DOMAIN_KEYWORDS.search(query):
        return True
    if _SITE_KEYWORDS.search(query) and context is not None:
        return True
    return False


CORS_ALLOW_ORIGINS = _get_csv_env("CORS_ALLOW_ORIGINS", DEFAULT_CORS_ALLOW_ORIGINS)

# ── Global clients (initialized on startup) ────────────────────────────

openai_client: OpenAI | None = None
pinecone_client: Pinecone | None = None
active_index_name: str = PINECONE_INDEX  # mutable — switched via API
_template_store: dict = {}  # stores uploaded underwriting template bytes

# Per-agent tunable settings
agent_settings: dict = {
    "librarian": {
        "top_k": 3,
        "description": "Retrieves semantic blueprints from the ContextLibrary to guide the Writer.",
    },
    "researcher": {
        "top_k": 60,
        "temperature": 0.1,
        "description": "Queries the KnowledgeStore and synthesizes facts with citations.",
    },
    "writer": {
        "temperature": 0.1,
        "description": "Combines research with the blueprint to generate the final response.",
    },
    "summarizer": {
        "temperature": 0.1,
        "max_length": 2000,
        "description": "Condenses large outputs to stay within token limits.",
    },
}


@asynccontextmanager
async def lifespan(_app: FastAPI):
    global openai_client, pinecone_client, active_index_name

    openai_key = os.getenv("OPENAI_API_KEY", "")
    pinecone_key = os.getenv("PINECONE_API_KEY", "")

    if not openai_key:
        logging.error("OPENAI_API_KEY not set!")
    if not pinecone_key:
        logging.error("PINECONE_API_KEY not set!")

    openai_client = OpenAI(api_key=openai_key)
    pinecone_client = Pinecone(api_key=pinecone_key)

    # Verify Pinecone index exists; auto-switch if the configured one is gone
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        total_vectors = stats.get("total_vector_count", 0)
        logging.info(f"✅ Pinecone connected: index={active_index_name}, vectors={total_vectors}")
    except Exception as e:
        logging.warning(f"Configured index '{active_index_name}' not reachable: {e}")
        # Try to fall back to the first available index
        try:
            available = pinecone_client.list_indexes()
            if available:
                first = available[0].name
                active_index_name = first
                idx = pinecone_client.Index(active_index_name)
                stats = idx.describe_index_stats()
                total_vectors = stats.get("total_vector_count", 0)
                logging.info(f"✅ Auto-switched to index '{active_index_name}', vectors={total_vectors}")
            else:
                logging.warning("No Pinecone indexes available")
        except Exception as e2:
            logging.error(f"Failed to auto-switch index: {e2}")

    logging.info(f"✅ MAS Backend ready — model={GENERATION_MODEL}, embedding={EMBEDDING_MODEL}")
    yield


app = FastAPI(title="UAP 485-x NYC Development Expert API", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ALLOW_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Models ──────────────────────────────────────────────────────────────

class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    messages: list[ChatMessage]
    use_rag: bool = True


class ChatResponse(BaseModel):
    reply: str
    sources: list[dict]


class UnderwritingUpdatesRequest(BaseModel):
    updates: dict[str, dict[str, str | int | float]] = {}


def _get_index():
    return pinecone_client.Index(active_index_name)


def _get_active_property_context() -> PropertyContext | None:
    try:
        return fetch_property_context(_get_index(), NAMESPACE_PROPERTY)
    except Exception as exc:
        logging.warning(f"Failed to load active property context: {exc}")
        return None


# ── Agent Settings ──────────────────────────────────────────────────────

@app.get("/api/settings")
async def get_settings():
    """Return current agent settings."""
    return {"settings": agent_settings}


@app.put("/api/settings")
async def update_settings(req: dict):
    """Update agent settings. Accepts partial updates per agent."""
    ALLOWED_KEYS = {
        "librarian": {"top_k": (int, 1, 20)},
        "researcher": {"top_k": (int, 1, 100), "temperature": (float, 0.0, 2.0)},
        "writer": {"temperature": (float, 0.0, 2.0)},
        "summarizer": {"temperature": (float, 0.0, 2.0), "max_length": (int, 100, 10000)},
    }
    settings_input = req.get("settings", {})
    for agent_name, params in settings_input.items():
        if agent_name not in ALLOWED_KEYS:
            continue
        for key, value in params.items():
            if key not in ALLOWED_KEYS[agent_name]:
                continue
            expected_type, min_val, max_val = ALLOWED_KEYS[agent_name][key]
            try:
                casted = expected_type(value)
                casted = max(min_val, min(max_val, casted))
                agent_settings[agent_name][key] = casted
            except (ValueError, TypeError):
                continue
    logging.info(f"Agent settings updated: {agent_settings}")
    return {"settings": agent_settings}


# ── Pinecone Index Management ───────────────────────────────────────────

class CreateIndexRequest(BaseModel):
    name: str
    dimension: int = 3072
    metric: str = "cosine"
    cloud: str = "aws"
    region: str = "us-east-1"


class SwitchIndexRequest(BaseModel):
    name: str


@app.get("/api/indexes")
async def list_indexes():
    """List all Pinecone indexes in the account."""
    try:
        indexes = pinecone_client.list_indexes()
        result = []
        for idx_model in indexes:
            status = idx_model.status
            result.append({
                "name": idx_model.name,
                "dimension": idx_model.dimension,
                "metric": idx_model.metric,
                "host": idx_model.host,
                "ready": status["ready"] if status else False,
                "state": status["state"] if status else "Unknown",
            })
        return {"indexes": result, "active": active_index_name}
    except Exception as e:
        logging.error(f"Failed to list indexes: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/indexes/active")
async def get_active_index():
    """Return the currently active Pinecone index with stats."""
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        ns = stats.get("namespaces", {})
        return {
            "name": active_index_name,
            "total_vectors": stats.get("total_vector_count", 0),
            "dimension": stats.get("dimension", 0),
            "namespaces": {
                k: {"vector_count": v.get("vector_count", 0)}
                for k, v in ns.items()
            },
        }
    except Exception as e:
        logging.error(f"Failed to get active index stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/indexes")
async def create_index(req: CreateIndexRequest):
    """Create a new Pinecone serverless index."""
    try:
        pinecone_client.create_index(
            name=req.name,
            dimension=req.dimension,
            metric=req.metric,
            spec=ServerlessSpec(cloud=req.cloud, region=req.region),
        )
        logging.info(f"Created Pinecone index: {req.name}")
        return {"created": req.name}
    except Exception as e:
        logging.error(f"Failed to create index: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/indexes/{index_name}")
async def delete_index(index_name: str):
    """Delete a Pinecone index."""
    if index_name == active_index_name:
        raise HTTPException(status_code=400, detail="Cannot delete the currently active index. Switch to another first.")
    try:
        pinecone_client.delete_index(name=index_name)
        logging.info(f"Deleted Pinecone index: {index_name}")
        return {"deleted": index_name}
    except Exception as e:
        logging.error(f"Failed to delete index: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/indexes/switch")
async def switch_index(req: SwitchIndexRequest):
    """Switch the active Pinecone index used for chat and uploads."""
    global active_index_name
    # Verify the index exists and is ready
    try:
        info = pinecone_client.describe_index(req.name)
        status = info.status
        if not status["ready"]:
            raise HTTPException(status_code=400, detail=f"Index '{req.name}' is not ready")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"Index '{req.name}' not found: {e}")

    active_index_name = req.name
    logging.info(f"Switched active index to: {active_index_name}")
    return {"active": active_index_name}


# ── Live Property Context ───────────────────────────────────────────────

@app.get("/api/property/search-address", response_model=PropertySearchResponse)
async def search_property_address(q: str = Query("", description="NYC address or 10-digit BBL")):
    try:
        results = await property_service.search_address(q)
        return PropertySearchResponse(results=results, query=q)
    except Exception as exc:
        logging.error(f"Property address search failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Address search failed: {exc}")


@app.get("/api/property/validate-lot", response_model=ValidatedLotInfo)
async def validate_property_lot(bbl: str = Query(..., description="10-digit BBL")):
    try:
        result = await property_service.validate_lot(bbl)
        return ValidatedLotInfo(**result)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logging.error(f"Property lot validation failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Lot validation failed: {exc}")


@app.get("/api/property/block-lots", response_model=BlockLotsResponse)
async def get_property_block_lots(
    borough: int = Query(..., ge=1, le=5, description="Borough code (1-5)"),
    block: int = Query(..., gt=0, description="Tax block"),
):
    try:
        result = await property_service.get_block_lots(borough, block)
        return BlockLotsResponse(**result)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logging.error(f"Property block lookup failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Block lot lookup failed: {exc}")


@app.put("/api/property/context", response_model=PropertyContext)
async def set_property_context(req: PropertyContextRequest):
    try:
        context = await property_service.build_property_context(req.primary_bbl, req.adjacent_bbls)
        embedding = get_embedding(context.property_brief, client=openai_client, embedding_model=EMBEDDING_MODEL)
        upsert_property_context(_get_index(), NAMESPACE_PROPERTY, embedding, context)
        logging.info(f"Stored active property context for index '{active_index_name}': {context.primary_bbl}")
        return context
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except LookupError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except Exception as exc:
        logging.error(f"Set property context failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Property context update failed: {exc}")


@app.get("/api/property/context", response_model=PropertyContext | None)
async def get_property_context():
    try:
        return _get_active_property_context()
    except Exception as exc:
        logging.error(f"Get property context failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Property context lookup failed: {exc}")


@app.delete("/api/property/context")
async def clear_property_context():
    try:
        delete_property_context(_get_index(), NAMESPACE_PROPERTY)
        logging.info(f"Cleared active property context for index '{active_index_name}'")
        return {"cleared": True}
    except Exception as exc:
        logging.error(f"Clear property context failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Property context clear failed: {exc}")


# ── Blueprint Management (ContextLibrary) ───────────────────────────────

class CreateBlueprintRequest(BaseModel):
    subject: str
    instructions: str


@app.get("/api/blueprints")
async def list_blueprints():
    """List all blueprints in the ContextLibrary namespace."""
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        dim = stats.get("dimension", 3072) or 3072
        # Fetch all vectors in ContextLibrary using a zero-vector query
        # (Pinecone doesn't have a "list" — we query with a dummy and high top_k)
        dummy_vec = [0.0] * int(dim)
        results = idx.query(
            vector=dummy_vec,
            top_k=100,
            namespace=NAMESPACE_CONTEXT,
            include_metadata=True,
        )
        blueprints = []
        for m in results.get("matches", []):
            meta = m.get("metadata", {})
            blueprints.append({
                "id": m["id"],
                "subject": meta.get("subject", "Unknown"),
                "instructions": meta.get("text", ""),
            })
        return {"blueprints": blueprints}
    except Exception as e:
        logging.error(f"Failed to list blueprints: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/blueprints")
async def create_blueprint(req: CreateBlueprintRequest):
    """Create a new blueprint in the ContextLibrary namespace."""
    if not req.subject.strip() or not req.instructions.strip():
        raise HTTPException(status_code=400, detail="Subject and instructions are required")

    try:
        idx = pinecone_client.Index(active_index_name)
        # Embed the subject so the Librarian can match it semantically
        embedding = get_embedding(req.subject.strip(), client=openai_client, embedding_model=EMBEDDING_MODEL)
        vec_id = f"blueprint__{req.subject.strip().lower().replace(' ', '_')}"
        idx.upsert(
            vectors=[{
                "id": vec_id,
                "values": embedding,
                "metadata": {
                    "text": req.instructions.strip(),
                    "subject": req.subject.strip(),
                    "source": "blueprint",
                },
            }],
            namespace=NAMESPACE_CONTEXT,
        )
        logging.info(f"Created blueprint: {req.subject}")
        return {"id": vec_id, "subject": req.subject.strip()}
    except Exception as e:
        logging.error(f"Failed to create blueprint: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class GenerateBlueprintRequest(BaseModel):
    subject: str


@app.post("/api/blueprints/generate")
async def generate_blueprint(req: GenerateBlueprintRequest):
    """Use AI to generate blueprint instructions for a subject, then store it."""
    subject = req.subject.strip()
    if not subject:
        raise HTTPException(status_code=400, detail="Subject is required")

    try:
        # Generate instructions via GPT
        system_prompt = (
            "You are writing blueprint instructions for an AI Writer that is exclusively focused on "
            "NYC UAP / 485-x building development strategy. Given a subject domain, produce instructions "
            "for how the Writer should format, structure, and tone responses inside that NYC development context.\n\n"
            "Include: developer-first tone, response structure, terminology guidance, profitability framing, "
            "assumption handling, and how to cite source-grounded constraints.\n\n"
            "Be specific and practical. Output ONLY the instructions, no preamble."
        )
        response = openai_client.chat.completions.create(
            model=GENERATION_MODEL,
            temperature=0.4,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Generate blueprint instructions for the subject: {subject}"},
            ],
        )
        instructions = response.choices[0].message.content.strip()

        # Upsert into Pinecone ContextLibrary
        idx = pinecone_client.Index(active_index_name)
        embedding = get_embedding(subject, client=openai_client, embedding_model=EMBEDDING_MODEL)
        vec_id = f"blueprint__{subject.lower().replace(' ', '_')}"
        idx.upsert(
            vectors=[{
                "id": vec_id,
                "values": embedding,
                "metadata": {
                    "text": instructions,
                    "subject": subject,
                    "source": "blueprint",
                },
            }],
            namespace=NAMESPACE_CONTEXT,
        )
        logging.info(f"AI-generated blueprint for '{subject}'")
        return {"id": vec_id, "subject": subject, "instructions": instructions}
    except Exception as e:
        logging.error(f"Failed to generate blueprint: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/blueprints/{blueprint_id:path}")
async def delete_blueprint(blueprint_id: str):
    """Delete a blueprint from the ContextLibrary namespace."""
    try:
        idx = pinecone_client.Index(active_index_name)
        idx.delete(ids=[blueprint_id], namespace=NAMESPACE_CONTEXT)
        logging.info(f"Deleted blueprint: {blueprint_id}")
        return {"deleted": blueprint_id}
    except Exception as e:
        logging.error(f"Failed to delete blueprint: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Routes ──────────────────────────────────────────────────────────────

@app.get("/api/health")
async def health():
    try:
        idx = pinecone_client.Index(active_index_name)
        stats = idx.describe_index_stats()
        total = stats.get("total_vector_count", 0)
    except Exception:
        total = 0
    return {"status": "ok", "documents": total, "active_index": active_index_name}


@app.post("/api/chat", response_model=ChatResponse)
async def chat(req: ChatRequest):
    # Extract the last user message as the goal
    last_user_msg = ""
    for m in reversed(req.messages):
        if m.role == "user":
            last_user_msg = m.content
            break

    if not last_user_msg:
        raise HTTPException(status_code=400, detail="No user message found")

    # Sanitize and moderate input
    try:
        sanitized = helper_sanitize_input(last_user_msg)
        helper_moderate_content(sanitized, openai_client)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    property_context = _get_active_property_context()

    # Run the Multi-Agent System pipeline
    try:
        result, trace = context_engine(
            goal=sanitized,
            client=openai_client,
            pc=pinecone_client,
            index_name=active_index_name,
            generation_model=GENERATION_MODEL,
            embedding_model=EMBEDDING_MODEL,
            namespace_context=NAMESPACE_CONTEXT,
            namespace_knowledge=NAMESPACE_KNOWLEDGE,
            agent_settings=agent_settings,
            property_context=property_context.model_dump() if property_context else None,
        )
    except Exception as e:
        logging.error(f"Context engine error: {e}")
        raise HTTPException(status_code=500, detail="Engine processing failed")

    if result is None:
        return ChatResponse(
            reply=f"I wasn't able to find a complete answer. Trace: {trace.status}",
            sources=[],
        )

    # Extract final text and sources from the MAS output
    if isinstance(result, str):
        reply_text = result
        sources = []
    elif isinstance(result, dict):
        reply_text = result.get("answer_with_sources", result.get("summary", str(result)))
        sources = result.get("sources", [])
    else:
        reply_text = str(result)
        sources = []

    # Format sources for the frontend contract
    formatted_sources = []
    for s in sources[:10]:
        if isinstance(s, dict):
            source_name = s.get("source", "Pinecone")
            formatted_sources.append({
                "filename": source_name,
                "distance": round(1 - s.get("score", 0), 4),
                "source_type": "property" if str(source_name).startswith("Active Property Context") else "document",
            })

    return ChatResponse(reply=reply_text, sources=formatted_sources)


@app.post("/api/upload")
async def upload_document(file: UploadFile = File(...)):
    """Upload any document and upsert its chunks into Pinecone KnowledgeStore."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    ext = os.path.splitext(file.filename)[1].lower()

    # Extract text based on file type
    text = _extract_text(content, ext, file.filename)

    if len(text.strip()) == 0:
        raise HTTPException(status_code=400, detail="File is empty or could not be parsed")

    # Chunk the text
    chunks = _chunk_text(text, chunk_size=800, overlap=200)
    if not chunks:
        raise HTTPException(status_code=400, detail="No chunks generated")

    # Upsert into Pinecone
    idx = pinecone_client.Index(active_index_name)
    embeddings = get_embeddings_batch(chunks, client=openai_client, embedding_model=EMBEDDING_MODEL)
    vectors = []
    for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
        vec_id = f"{file.filename}__chunk_{i}"
        vectors.append({
            "id": vec_id,
            "values": embedding,
            "metadata": {"text": chunk, "source": file.filename, "chunk_index": i},
        })

    # Upsert in batches of 100
    for batch_start in range(0, len(vectors), 100):
        batch = vectors[batch_start:batch_start + 100]
        idx.upsert(vectors=batch, namespace=NAMESPACE_KNOWLEDGE)

    logging.info(f"Uploaded {len(chunks)} chunks for '{file.filename}' to Pinecone.")
    return {"filename": file.filename, "chunks": len(chunks)}


@app.get("/api/documents")
async def list_documents():
    """List uploaded documents (by filename) from the active index's KnowledgeStore."""
    try:
        idx = pinecone_client.Index(active_index_name)
        filenames: dict[str, int] = {}  # filename → chunk count

        for page in idx.list(namespace=NAMESPACE_KNOWLEDGE):
            for item in page:
                vid = item if isinstance(item, str) else getattr(item, "id", str(item))
                if "__chunk_" in vid:
                    fname = vid.rsplit("__chunk_", 1)[0]
                    filenames[fname] = filenames.get(fname, 0) + 1

        docs = [{"filename": f, "chunks": c} for f, c in sorted(filenames.items())]
        total = sum(filenames.values())
        return {"documents": docs, "total_chunks": total}
    except Exception as e:
        logging.error(f"List documents failed: {e}")
        return {"documents": [], "total_chunks": 0}


@app.delete("/api/documents/{filename:path}")
async def delete_document(filename: str):
    """Delete all vectors for a given source filename from Pinecone."""
    try:
        idx = pinecone_client.Index(active_index_name)
        # Delete by metadata filter
        idx.delete(
            filter={"source": {"$eq": filename}},
            namespace=NAMESPACE_KNOWLEDGE,
        )
        logging.info(f"Deleted vectors for '{filename}' from Pinecone.")
        return {"deleted_chunks": 1}  # Pinecone doesn't return count
    except Exception as e:
        logging.error(f"Delete failed: {e}")
        raise HTTPException(status_code=500, detail=f"Delete failed: {e}")


def _extract_text(content: bytes, ext: str, filename: str) -> str:
    """Extract plain text from any supported file type."""
    # PDF — prefer pdfplumber for table-aware extraction, fall back to pypdf
    if ext == ".pdf":
        try:
            import pdfplumber
            pages_text: list[str] = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    # Extract tables first so we can represent them structurally
                    tables = page.extract_tables() or []
                    table_texts: list[str] = []
                    for table in tables:
                        if not table:
                            continue
                        header = table[0]
                        for row in table[1:]:
                            if not row:
                                continue
                            row_parts: list[str] = []
                            for idx, cell in enumerate(row):
                                col_name = header[idx] if header and idx < len(header) and header[idx] else f"Col{idx+1}"
                                cell_val = (cell or "").strip()
                                if cell_val:
                                    row_parts.append(f"{col_name}: {cell_val}")
                            if row_parts:
                                table_texts.append(" | ".join(row_parts))

                    # Get non-table text
                    page_text = page.extract_text() or ""
                    if table_texts:
                        page_text += "\n\n[Table data]\n" + "\n".join(table_texts)
                    pages_text.append(page_text)
            return "\n\n".join(pages_text)
        except ImportError:
            pass  # fall through to pypdf
        except Exception as e:
            logging.warning(f"pdfplumber failed for {filename}, falling back to pypdf: {e}")

        # Fallback: pypdf
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content))
            pages = [page.extract_text() or "" for page in reader.pages]
            return "\n\n".join(pages)
        except Exception as e:
            logging.error(f"PDF parse failed for {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse PDF: {e}")

    # DOCX
    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception as e:
            logging.error(f"DOCX parse failed for {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse DOCX: {e}")

    # XLSX / XLS — structured extraction preserving column headers per row
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"=== Sheet: {sheet} ===")
                header_cells: list[str] = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx == 0:
                        header_cells = [str(c).strip() if c is not None else f"Col{i+1}" for i, c in enumerate(row)]
                        lines.append("Columns: " + " | ".join(header_cells))
                        continue
                    parts: list[str] = []
                    for col_idx, cell in enumerate(row):
                        if cell is None:
                            continue
                        col_name = header_cells[col_idx] if col_idx < len(header_cells) else f"Col{col_idx+1}"
                        parts.append(f"{col_name}: {cell}")
                    if parts:
                        lines.append(" | ".join(parts))
            return "\n".join(lines)
        except Exception as e:
            logging.error(f"Excel parse failed for {filename}: {e}")
            raise HTTPException(status_code=400, detail=f"Failed to parse Excel: {e}")

    # Everything else — treat as UTF-8 text
    return content.decode("utf-8", errors="replace")


def _chunk_text(text: str, chunk_size: int = 800, overlap: int = 200) -> list[str]:
    """Split text into overlapping chunks, preferring paragraph boundaries."""
    import re
    # Split into paragraphs first (double newlines, section headers, page breaks)
    paragraphs = re.split(r'\n{2,}|(?=^={3,}|^---)', text, flags=re.MULTILINE)
    paragraphs = [p.strip() for p in paragraphs if p.strip()]

    chunks: list[str] = []
    current_chunk = ""

    for para in paragraphs:
        # If a single paragraph exceeds chunk_size, sub-split it the old way
        if len(para) > chunk_size:
            # Flush what we have
            if current_chunk.strip():
                chunks.append(current_chunk.strip())
                current_chunk = ""
            start = 0
            while start < len(para):
                end = start + chunk_size
                # Try to break at a sentence or line boundary
                if end < len(para):
                    boundary = max(
                        para.rfind(". ", start + chunk_size // 2, end),
                        para.rfind("\n", start + chunk_size // 2, end),
                    )
                    if boundary > start:
                        end = boundary + 1
                chunks.append(para[start:end].strip())
                start = end - overlap if end - overlap > start else end
            continue

        # Would adding this paragraph exceed chunk_size?
        candidate = (current_chunk + "\n\n" + para).strip() if current_chunk else para
        if len(candidate) <= chunk_size:
            current_chunk = candidate
        else:
            if current_chunk.strip():
                chunks.append(current_chunk.strip())
            # Start new chunk with overlap from the end of previous chunk
            if overlap > 0 and chunks:
                tail = chunks[-1][-overlap:]
                current_chunk = tail + "\n\n" + para
            else:
                current_chunk = para

    if current_chunk.strip():
        chunks.append(current_chunk.strip())

    return [c for c in chunks if c.strip()]


def _select_diversified_source_chunks(
    matches: list[dict],
    *,
    max_sources: int = 10,
    max_chunks_per_source: int = 8,
    max_total_chunks: int = 40,
) -> tuple[list[str], dict[str, str]]:
    """Balance retrieved chunks across documents so one file does not dominate extraction."""
    chunks_by_source: dict[str, list[tuple[int | None, str]]] = {}
    source_order: list[str] = []

    for match in matches:
        metadata = match.get("metadata", {}) or {}
        chunk_text = metadata.get("text", "")
        source_name = str(metadata.get("source", "Unknown source")).strip() or "Unknown source"
        if not chunk_text:
            continue

        if source_name not in chunks_by_source:
            chunks_by_source[source_name] = []
            source_order.append(source_name)

        if len(chunks_by_source[source_name]) >= max_chunks_per_source:
            continue

        raw_chunk_index = metadata.get("chunk_index")
        chunk_index = raw_chunk_index if isinstance(raw_chunk_index, int) else None
        chunks_by_source[source_name].append((chunk_index, chunk_text))

    selected_sources = source_order[:max_sources]
    source_chunks: list[str] = []
    source_name_lookup: dict[str, str] = {}

    round_index = 0
    while len(source_chunks) < max_total_chunks:
        added_any = False
        for source_name in selected_sources:
            chunks = chunks_by_source.get(source_name, [])
            if round_index >= len(chunks):
                continue

            chunk_index, chunk_text = chunks[round_index]
            header = f"[Source: {source_name}]"
            if chunk_index is not None:
                header += f"\n[Chunk {chunk_index}]"
            source_chunks.append(f"{header}\n{chunk_text}")
            source_name_lookup[source_name.lower()] = source_name
            added_any = True

            if len(source_chunks) >= max_total_chunks:
                break

        if not added_any:
            break
        round_index += 1

    return source_chunks, source_name_lookup


def _pick_relevant_glossary_sections(labels: list[str], sheet_name: str) -> str:
    """Select only the glossary sections relevant to this sheet's labels."""
    from underwriting_domain import UNDERWRITING_GLOSSARY

    sheet_lower = sheet_name.lower()
    label_blob = " ".join(labels).lower()
    combined = f"{sheet_lower} {label_blob}"

    section_keywords: dict[str, list[str]] = {
        "Property Information": ["address", "bbl", "block", "lot", "zoning", "far", "buildable", "height", "borough", "overlay"],
        "Unit Mix": ["unit", "studio", "1br", "2br", "3br", "bedroom", "ami", "affordable", "market rate", "dwelling", "duf"],
        "Revenue": ["rent", "gpr", "egi", "vacancy", "income", "laundry", "parking", "storage", "commercial", "collection"],
        "Operating Expenses": ["expense", "opex", "tax", "insurance", "payroll", "management", "repair", "utility", "water", "electric", "gas", "fuel", "elevator", "r&m", "reserve"],
        "NOI & Valuation": ["noi", "cap rate", "valuation", "appraised", "price per", "grm"],
        "Acquisition & Financing": ["purchase", "acquisition", "closing", "transfer", "hard cost", "soft cost", "tdc", "equity", "ltv", "ltc"],
        "Debt Service": ["mortgage", "loan", "interest", "amortization", "debt service", "dscr", "construction loan", "permanent", "mezzanine"],
        "Returns": ["cash flow", "irr", "cash-on-cash", "coc", "equity multiple", "hold period", "exit cap", "reversion", "roi", "yield", "btcf", "atcf"],
        "Tax Programs": ["uap", "485", "421", "abatement", "icap", "j-51", "prevailing", "pilot"],
        "Development": ["gsf", "nsf", "rsf", "efficiency", "stories", "floor", "cellar", "parking", "construction type", "lease-up", "stabilization"],
        "Rent Regulation": ["dhcr", "rgb", "mci", "iai", "hstpa", "stabilized", "regulated", "decontrol"],
        "Sources & Uses": ["source", "uses", "land cost", "contingency", "developer fee", "lihtc", "hpd subsidy", "hdc bond", "gap funding", "interest reserve"],
        "Pro Forma Projections": ["year 1", "year 2", "pro forma", "rent growth", "expense growth", "npv", "discount rate", "terminal", "levered irr"],
        "Sensitivity Analysis": ["base case", "downside", "upside", "break-even", "stress", "sensitivity"],
        "Deal Structure": ["gp", "lp", "sponsor", "preferred return", "promote", "waterfall", "catch-up", "capital stack", "joint venture"],
        "485-x Program": ["485-x", "benefit period", "phase-out", "affordability lock", "prevailing wage", "regulatory agreement", "hpd marketing"],
        "UAP Scenarios": ["as-of-right", "full bonus", "avoid prevailing", "avoid 40%", "bonus floor", "optimized"],
        "NYC Compliance & Expenses": ["scrie", "drie", "hpd violation", "lead paint", "local law", "ll97", "ll11", "ll87", "landmark", "certificate of occupancy", "dob", "sro"],
        "Abbreviations": [],
    }

    selected_sections: list[str] = []
    for section, keywords in section_keywords.items():
        if section == "Abbreviations":
            selected_sections.append(section)
            continue
        if any(kw in combined for kw in keywords):
            selected_sections.append(section)

    for core in ["Property Information", "Unit Mix", "Revenue", "Operating Expenses", "NOI & Valuation"]:
        if core not in selected_sections:
            selected_sections.append(core)

    lines: list[str] = ["DOMAIN GLOSSARY (relevant sections):\n"]
    for section in selected_sections:
        entries = UNDERWRITING_GLOSSARY.get(section, [])
        if not entries:
            continue
        lines.append(f"## {section}")
        for label, meaning in entries:
            lines.append(f"  - {label}: {meaning}")
        lines.append("")
    return "\n".join(lines)


# ── Underwriting Template ───────────────────────────────────────────────

def _safe_cell_value(val):
    """Convert cell value to a JSON-serializable type."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        return round(val, 2)
    if isinstance(val, int):
        return val
    if isinstance(val, str):
        return val
    return str(val)


def _col_letter(col: int) -> str:
    """Convert 1-based column number to Excel column letter(s)."""
    result = ""
    while col > 0:
        col -= 1
        result = chr(65 + col % 26) + result
        col //= 26
    return result


@app.post("/api/underwriting/parse-template")
async def parse_underwriting_template(file: UploadFile = File(...)):
    """Upload an Excel underwriting template and return its parsed structure."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    content = await file.read()
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx / .xls) are supported")

    import openpyxl

    wb_vals = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    wb_fmls = openpyxl.load_workbook(io.BytesIO(content), data_only=False)

    sheets = []
    formula_refs_by_sheet: dict[str, list[str]] = {}
    for name in wb_vals.sheetnames:
        ws_v = wb_vals[name]
        ws_f = wb_fmls[name]
        max_r = ws_v.max_row or 0
        max_c = ws_v.max_column or 0

        data = []
        formula_refs: list[str] = []
        for r in range(1, max_r + 1):
            row = []
            for c in range(1, max_c + 1):
                value_cell = ws_v.cell(r, c)
                formula_cell = ws_f.cell(r, c)
                fml = formula_cell.value
                is_formula = isinstance(fml, str) and fml.startswith("=")
                cell = build_underwriting_cell_payload(
                    value_cell.value,
                    row=r,
                    col=c,
                    is_formula=is_formula,
                    number_format=formula_cell.number_format,
                    epoch=wb_vals.epoch,
                )
                if cell is None:
                    row.append(None)
                else:
                    if is_formula:
                        formula_refs.append(f"{_col_letter(c)}{r}")
                    row.append(cell)
            data.append(row)

        sheets.append({"name": name, "data": data, "maxRow": max_r, "maxCol": max_c})
        formula_refs_by_sheet[name] = formula_refs

    calc_context = build_underwriting_calculation_context(
        file.filename,
        content,
        formula_refs_by_sheet,
    )
    _template_store["current"] = {
        "filename": file.filename,
        "bytes": content,
        "formula_refs_by_sheet": formula_refs_by_sheet,
        "calc_context": calc_context,
    }
    logging.info(f"Parsed underwriting template: {file.filename} ({len(sheets)} sheets)")
    return {"filename": file.filename, "sheets": sheets}


@app.post("/api/underwriting/extract")
async def extract_underwriting_values():
    """Use RAG to extract values from uploaded source documents for each template sheet."""
    if "current" not in _template_store:
        raise HTTPException(status_code=400, detail="No template uploaded yet")

    import openpyxl

    # Check if there are documents to extract from
    idx = _get_index()
    stats = idx.describe_index_stats()
    ns = stats.get("namespaces", {})
    knowledge_count = ns.get(NAMESPACE_KNOWLEDGE, {}).get("vector_count", 0)
    if knowledge_count == 0:
        return {"updates": {}, "message": "No documents uploaded. Upload source documents first."}

    wb = openpyxl.load_workbook(io.BytesIO(_template_store["current"]["bytes"]), data_only=True)
    wb_f = openpyxl.load_workbook(io.BytesIO(_template_store["current"]["bytes"]), data_only=False)

    all_updates: dict[str, dict] = {}
    all_sources: dict[str, dict[str, str]] = {}
    all_confidence: dict[str, dict[str, str]] = {}

    ROWS_PER_BATCH = 40

    for name in wb.sheetnames:
        name_lower = name.lower()
        if "(auto)" in name_lower or name_lower.endswith("auto") or " auto" in name_lower:
            logging.info(f"  ⏭ Skipping sheet '{name}' (auto-calculated)")
            continue

        ws = wb[name]
        ws_f_s = wb_f[name]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0

        if max_r == 0 or max_c == 0:
            logging.info(f"  ⏭ Skipping sheet '{name}' (empty)")
            continue

        labels: list[str] = []

        # Build a grid: grid[r][c] → (display_value, is_formula, is_empty)
        grid: dict[int, dict[int, tuple]] = {}
        for r in range(1, max_r + 1):
            grid[r] = {}
            for c in range(1, max_c + 1):
                val = ws.cell(r, c).value
                fml = ws_f_s.cell(r, c).value
                is_formula = isinstance(fml, str) and fml.startswith("=")
                if is_formula:
                    grid[r][c] = (None, True, False)
                elif val is None:
                    grid[r][c] = (None, False, True)
                else:
                    safe = _safe_cell_value(val)
                    grid[r][c] = (safe, False, False)
                    if isinstance(val, str) and not val.replace(".", "").replace("-", "").replace(",", "").replace(" ", "").isdigit():
                        labels.append(val)

        # Detect header row — scan rows 1-3 for the one with the most text cells
        header_row: dict[int, str] = {}
        header_row_num = 0
        best_text_count = 0
        for candidate_r in range(1, min(max_r + 1, 4)):
            text_count = 0
            candidate_headers: dict[int, str] = {}
            for c in range(1, max_c + 1):
                val, is_f, is_e = grid[candidate_r].get(c, (None, False, True))
                if isinstance(val, str) and val.strip():
                    text_count += 1
                    candidate_headers[c] = val.strip()
            if text_count > best_text_count:
                best_text_count = text_count
                header_row = candidate_headers
                header_row_num = candidate_r

        # Build row-by-row descriptions
        row_descriptions: list[str] = []
        header_line = ""
        if header_row:
            hdr_parts = []
            for c in range(1, max_c + 1):
                col_letter = _col_letter(c)
                hdr_val = header_row.get(c, "")
                hdr_parts.append(f"{col_letter}=\"{hdr_val}\"" if hdr_val else col_letter)
            header_line = f"Row {header_row_num} (header): {' | '.join(hdr_parts)}"
            row_descriptions.append(header_line)

        start_row = (header_row_num + 1) if header_row else 1
        for r in range(start_row, max_r + 1):
            parts: list[str] = []
            row_has_content = False
            row_label = None
            for c in range(1, min(max_c + 1, 4)):
                val, is_f, is_e = grid[r].get(c, (None, False, True))
                if isinstance(val, str) and val.strip():
                    row_label = val.strip()
                    break

            for c in range(1, max_c + 1):
                val, is_formula, is_empty = grid[r].get(c, (None, False, True))
                col_letter = _col_letter(c)
                coord = f"{col_letter}{r}"
                col_header = header_row.get(c, "")

                if is_formula:
                    parts.append(f"{coord}=[formula]")
                    row_has_content = True
                elif is_empty:
                    context_hint = f" ({col_header})" if col_header else ""
                    parts.append(f"{coord}=<empty>{context_hint}")
                else:
                    # Mark existing values with [current] so LLM knows to overwrite
                    if isinstance(val, (int, float)):
                        parts.append(f"{coord}={val} [current]")
                    else:
                        parts.append(f"{coord}={val}")
                    row_has_content = True

            if row_has_content or any(
                not grid[r].get(c, (None, False, True))[2]
                for c in range(1, max_c + 1)
            ):
                label_hint = f" [{row_label}]" if row_label else ""
                row_descriptions.append(f"Row {r}{label_hint}: {' | '.join(parts)}")

        if not row_descriptions:
            logging.info(f"  ⏭ Skipping sheet '{name}' (no content rows)")
            continue

        logging.info(f"  📊 Sheet '{name}': {len(row_descriptions)} rows, {len(labels)} labels")

        # ── RAG retrieval ──────────────────────────────────────────────
        LABELS_PER_QUERY = 8
        label_groups = [labels[i:i + LABELS_PER_QUERY] for i in range(0, max(len(labels), 1), LABELS_PER_QUERY)]
        all_matches: list[dict] = []
        seen_ids: set[str] = set()

        # Batch-embed all label-group queries in one API call
        query_texts = [
            (f"UAP underwriting {name}: " + " ".join(group)) if group else f"UAP underwriting {name}"
            for group in label_groups
        ]
        query_embeddings = get_embeddings_batch(query_texts, client=openai_client, embedding_model=EMBEDDING_MODEL)

        for query_embedding in query_embeddings:
            results = idx.query(
                vector=query_embedding,
                top_k=50,
                namespace=NAMESPACE_KNOWLEDGE,
                include_metadata=True,
            )
            for m in results.get("matches", []):
                mid = m.get("id", "")
                if mid not in seen_ids:
                    seen_ids.add(mid)
                    all_matches.append(m)

        all_matches.sort(key=lambda m: m.get("score", 0), reverse=True)

        source_chunks, source_name_lookup = _select_diversified_source_chunks(all_matches)

        if not source_chunks:
            logging.info(f"  ⏭ Skipping sheet '{name}' (no source chunks)")
            continue

        logging.info(f"  📄 Sheet '{name}': {len(all_matches)} RAG matches → {len(source_chunks)} chunks from {len(source_name_lookup)} docs")

        context_text = "\n---\n".join(source_chunks)
        glossary_context = _pick_relevant_glossary_sections(labels, name)

        # ── Split rows into batches ────────────────────────────────────
        data_rows = [rd for rd in row_descriptions if not rd.startswith(f"Row {header_row_num} (header)")]
        batches = [data_rows[i:i + ROWS_PER_BATCH] for i in range(0, max(len(data_rows), 1), ROWS_PER_BATCH)]

        sheet_updates: dict[str, object] = {}
        sheet_sources: dict[str, str] = {}
        sheet_confidence: dict[str, str] = {}

        logging.info(f"  🔄 Sheet '{name}': {len(batches)} batch(es) of ≤{ROWS_PER_BATCH} rows")

        for batch_idx, batch_rows in enumerate(batches):
            batch_desc = (header_line + "\n" if header_line else "") + "\n".join(batch_rows)

            prompt = (
                f'You are filling a UAP underwriting Excel spreadsheet from source documents.\n\n'
                f'Sheet: "{name}" (batch {batch_idx + 1}/{len(batches)})\n\n'
                f'LAYOUT KEY:\n'
                f'  CellRef=Value — cell with its current value\n'
                f'  CellRef=<empty> — cell with no value, needs to be filled\n'
                f'  CellRef=Value [current] — cell with an existing value that SHOULD BE OVERWRITTEN '
                f'if the source documents contain data for it\n'
                f'  CellRef=[formula] — auto-calculated, NEVER fill these\n'
                f'  Row labels in [brackets] describe each row\'s purpose\n'
                f'  The header row defines what each column means\n\n'
                f'Spreadsheet layout:\n{batch_desc}\n\n'
                f'Source document excerpts:\n{context_text}\n\n'
                f'YOUR TASK:\n'
                f'1. FILL EVERY POSSIBLE CELL.  Check every <empty> cell AND every [current] cell.\n'
                f'   For [current] cells, OVERWRITE with the value from the source documents —\n'
                f'   the source documents are the authority, not the existing spreadsheet values.\n'
                f'2. Return a JSON object mapping cell refs (e.g. "B5") to:\n'
                f'   {{"value": <number_or_string>, "source": "<exact filename>", "confidence": "high|medium|low"}}\n'
                f'3. Use the exact source name from the [Source: ...] header of each excerpt.\n'
                f'4. Numbers should be plain (no $ or commas).  Text fields should be strings.\n'
                f'5. NEVER fill [formula] cells.\n'
                f'6. Confidence: "high" = explicitly stated in source, "medium" = reasonable inference, '
                f'   "low" = educated guess or industry standard assumption.\n'
                f'7. BE AGGRESSIVE — fill as many cells as possible.  It is better to fill a cell with '
                f'   medium/low confidence than to leave it empty.\n'
                f'8. NEVER fill a numeric cell with 0 as a placeholder or guess.  Only use 0 if the '
                f'   source document explicitly states the value is zero.'
            )

            sys_msg = (
                "You are an expert NYC real estate underwriting analyst.  You read source documents "
                "(rent rolls, T-12 operating statements, appraisals, offering memorandums, tax bills, "
                "surveys, environmental reports, financial projections) and populate underwriting "
                "spreadsheet cells.  The SOURCE DOCUMENTS are the single source of truth — if a cell "
                "already has a value marked [current], you MUST overwrite it with the value from the "
                "source documents.  Use column headers and row labels to determine which value belongs "
                "in which cell.  Return only valid JSON.\n\n" + glossary_context
            )

            try:
                response = openai_client.chat.completions.create(
                    model=GENERATION_MODEL,
                    temperature=0.15,
                    messages=[
                        {"role": "system", "content": sys_msg},
                        {"role": "user", "content": prompt},
                    ],
                    response_format={"type": "json_object"},
                )
                result_text = response.choices[0].message.content.strip()
                updates = json.loads(result_text)
                if isinstance(updates, dict):
                    batch_count = 0
                    for raw_ref, payload in updates.items():
                        if not isinstance(raw_ref, str):
                            continue
                        ref = raw_ref.strip().upper()
                        value = payload
                        source_name = None
                        confidence = None
                        if isinstance(payload, dict):
                            value = payload.get("value")
                            raw_source_name = payload.get("source")
                            if isinstance(raw_source_name, str):
                                source_name = source_name_lookup.get(raw_source_name.strip().lower())
                            raw_confidence = payload.get("confidence")
                            if isinstance(raw_confidence, str) and raw_confidence.lower() in ("high", "medium", "low"):
                                confidence = raw_confidence.lower()

                        if not isinstance(value, (str, int, float, bool)) and value is not None:
                            continue

                        # Discard zero-value low-confidence fills — likely hallucinated placeholders
                        if value == 0 and (confidence or "medium") == "low":
                            continue

                        sheet_updates[ref] = value
                        sheet_confidence[ref] = confidence or "medium"
                        if source_name:
                            sheet_sources[ref] = source_name
                        elif len(source_name_lookup) == 1:
                            sheet_sources[ref] = next(iter(source_name_lookup.values()))
                        batch_count += 1

                    logging.info(f"    ✅ Batch {batch_idx + 1}/{len(batches)}: {batch_count} cells")
            except Exception as e:
                logging.warning(f"    ❌ Batch {batch_idx + 1}/{len(batches)} failed for '{name}': {e}")

        if sheet_updates:
            all_updates[name] = sheet_updates
        if sheet_sources:
            all_sources[name] = sheet_sources
        if sheet_confidence:
            all_confidence[name] = sheet_confidence
        logging.info(f"  📊 Sheet '{name}' total: {len(sheet_updates)} cells extracted")

    total_cells = sum(len(v) for v in all_updates.values())
    logging.info(f"🏁 RAG extraction complete: {total_cells} cells across {len(all_updates)} sheets")
    return {"updates": all_updates, "sources": all_sources, "confidence": all_confidence}


@app.post("/api/underwriting/recalculate")
async def recalculate_underwriting_formulas(req: UnderwritingUpdatesRequest):
    """Recalculate workbook formulas using current cell updates without mutating the workbook."""
    if "current" not in _template_store:
        raise HTTPException(status_code=400, detail="No template uploaded")

    template = _template_store["current"]
    formula_values, warnings = calculate_underwriting_formula_values(
        template.get("calc_context"),
        req.updates,
    )
    return {
        "formulaValues": formula_values,
        "warnings": [warning.to_dict() for warning in warnings],
    }


@app.post("/api/underwriting/download")
async def download_filled_template(req: UnderwritingUpdatesRequest):
    """Apply cell updates to the stored template and return the filled .xlsx file."""
    if "current" not in _template_store:
        raise HTTPException(status_code=400, detail="No template uploaded")

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(_template_store["current"]["bytes"]))

    updates = req.updates
    for sheet_name, cells in updates.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for ref, value in cells.items():
            try:
                ws[ref] = value
            except Exception:
                continue

    enable_workbook_recalculation(wb)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    base = os.path.splitext(_template_store["current"]["filename"])[0]
    filename = f"{base}_filled.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "main:app",
        host=os.getenv("BACKEND_HOST", "0.0.0.0"),
        port=int(os.getenv("BACKEND_PORT", "8000")),
        reload=True,
    )
